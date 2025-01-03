# -*- coding: utf-8 -*-
from PySide6.QtCore import Qt, QDate
from PySide6.QtWidgets import QTableWidget, QTableWidgetItem, QMessageBox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import os
import json
from datetime import datetime

class DeliveryManager:
    def __init__(self, parent=None):
        self.parent = parent
        self.delivery_table = None
        self.stored_values = {}

    def setup_delivery_table(self, table):
        self.delivery_table = table
        self.setup_headers()
        
    def setup_headers(self):
        headers = ["Dátum"] + [f"Övezet {i}-{i+5}" for i in range(0, 45, 5)]
        self.delivery_table.setColumnCount(len(headers))
        self.delivery_table.setHorizontalHeaderLabels(headers)

    def handleM3Input(self):
        try:
            text = self.parent.m3_input.text().strip().replace(',', '.')
            if not text:
                QMessageBox.warning(self.parent, "Hiba", "Kérem adjon meg egy számot!")
                return
                
            value = float(text)
            if value < 0:
                QMessageBox.warning(self.parent, "Hiba", "Kérem pozitív számot adjon meg!")
                return
                
            self._store_m3_value(value)
            self._update_display()
            
        except ValueError:
            QMessageBox.warning(self.parent, "Hiba", "Kérem számot adjon meg (pl.: 6.0 vagy 6,0)")

    def _store_m3_value(self, value):
        current_date = self.parent.date_edit.date().toString('yyyy-MM-dd')
        current_zone = self.parent.km_combo.currentText()
        
        if current_date not in self.stored_values:
            self.stored_values[current_date] = {}
            
        if current_zone not in self.stored_values[current_date]:
            self.stored_values[current_date][current_zone] = []
            
        self.stored_values[current_date][current_zone].append(value)

    def _update_display(self):
        self.updateDeliveryTableWithStoredValues()
        self.parent.m3_input.clear()
        
        current_date = self.parent.date_edit.date().toString('yyyy-MM-dd')
        current_zone = self.parent.km_combo.currentText()
        self.updateM3Sum(current_date, current_zone)

    def updateM3Sum(self, current_date, current_zone):
        if current_date in self.stored_values and current_zone in self.stored_values[current_date]:
            values = self.stored_values[current_date][current_zone]
            values_text = " + ".join(f"{v:.1f}" for v in values)
            total = sum(values)
            self.parent.m3_sum.setText(f"({values_text}) = {total:.1f}")
        else:
            self.parent.m3_sum.setText("(0)")

    def getZoneColumn(self, zone_text):
        try:
            start_km = int(zone_text.split(' ')[1].split('-')[0])
            return (start_km // 5) + 1
        except:
            return 0

    def updateDeliveryTableWithStoredValues(self):
        current_date = self.parent.date_edit.date().toString('yyyy-MM-dd')
        if current_date in self.stored_values:
            for row in range(self.delivery_table.rowCount()):
                if self.delivery_table.item(row, 0) and self.delivery_table.item(row, 0).text() == current_date:
                    for zone, values in self.stored_values[current_date].items():
                        col = self.getZoneColumn(zone)
                        if col > 0:
                            display_text = " + ".join(f"{v:.1f}" for v in values)
                            new_item = QTableWidgetItem(display_text)
                            new_item.setTextAlignment(Qt.AlignCenter)
                            self.delivery_table.setItem(row, col, new_item)

    def saveDeliveryData(self):
        try:
            if not self.parent.driver_combo.currentText():
                QMessageBox.warning(self.parent, "Hiba", "Válasszon sofőrt!")
                return False

            current_driver = self.parent.driver_combo.currentText()
            month_dir = os.path.join('driver_records', current_driver, 
                                   f"{datetime.now().year}_{datetime.now().month:02d}")
            os.makedirs(month_dir, exist_ok=True)
            
            excel_path = os.path.join(month_dir, 'fuvar_nyilvantartas.xlsx')
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Fuvar adatok"

            # Headers
            headers = ["Dátum"] + [f"Övezet {i}-{i+5}" for i in range(0, 45, 5)]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            # Data
            for row in range(self.delivery_table.rowCount()):
                for col in range(self.delivery_table.columnCount()):
                    item = self.delivery_table.item(row, col)
                    cell = ws.cell(row=row + 2, column=col + 1)
                    cell.value = item.text() if item else ""
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

            wb.save(excel_path)
            QMessageBox.information(self.parent, "Siker", "Fuvar adatok mentve!")
            return True

        except Exception as e:
            QMessageBox.critical(self.parent, "Hiba", f"Mentési hiba: {str(e)}")
            return False

    def loadDeliveryData(self, current_driver):
        try:
            month_dir = os.path.join('driver_records', current_driver, 
                                   f"{datetime.now().year}_{datetime.now().month:02d}")
            excel_path = os.path.join(month_dir, 'fuvar_nyilvantartas.xlsx')
            
            if not os.path.exists(excel_path):
                return False

            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active

            for row in range(2, ws.max_row + 1):
                date = ws.cell(row=row, column=1).value
                if not date:
                    continue
                    
                for table_row in range(self.delivery_table.rowCount()):
                    table_date = self.delivery_table.item(table_row, 0)
                    if table_date and table_date.text() == str(date):
                        for col in range(1, self.delivery_table.columnCount()):
                            value = ws.cell(row=row, column=col+1).value
                            if value:
                                item = QTableWidgetItem(str(value))
                                item.setTextAlignment(Qt.AlignCenter)
                                self.delivery_table.setItem(table_row, col, item)
                        break

            return True

        except Exception as e:
            QMessageBox.warning(self.parent, "Hiba", f"Betöltési hiba: {str(e)}")
            return False