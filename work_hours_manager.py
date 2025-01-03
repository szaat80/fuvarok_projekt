# -*- coding: utf-8 -*-
from PySide6.QtWidgets import QTableWidget, QTableWidgetItem, QMessageBox
from PySide6.QtCore import Qt, QDate, QTime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os
from datetime import datetime

class WorkHoursManager:
    def __init__(self, parent=None):
        self.parent = parent
        self.work_table = None
        self.transport_table = None

    def setup_work_table(self, table):
        self.work_table = table# -*- coding: utf-8 -*-
from PySide6.QtWidgets import QTableWidget, QTableWidgetItem, QMessageBox
from PySide6.QtCore import Qt, QDate, QTime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os
from datetime import datetime

class WorkHoursManager:
    def __init__(self, parent=None):
        self.parent = parent
        self.work_table = None
        self.transport_table = None

    def setup_work_table(self, table):
        self.work_table = table
        self.setup_headers()

    def setup_transport_table(self, table):
        self.transport_table = table
        self.setup_transport_headers()

    def setup_headers(self):
        headers = [
            "Dátum", "Nap", 
            "Sima Munkanap\nKezdés", "Sima Munkanap\nVégzés", "Ledolgozott\nÓrák",
            "Műhely\nKezdés", "Műhely\nVégzés", "Műhely\nÓrák",
            "Szabadság", "Betegszabadság\n(TP)"
        ]
        self.work_table.setColumnCount(len(headers))
        self.work_table.setHorizontalHeaderLabels(headers)
    
        # Oszlopok szélességének beállítása
        for col in range(self.work_table.columnCount()):
            self.work_table.setColumnWidth(col, 120)

    def setup_transport_headers(self):
        headers = [
            "Dátum", "Fuvar ID", "Indulás", "Érkezés", "Távolság", "Üzemanyag"
        ]
        self.transport_table.setColumnCount(len(headers))
        self.transport_table.setHorizontalHeaderLabels(headers)
    
        # Oszlopok szélességének beállítása
        for col in range(self.transport_table.columnCount()):
            self.transport_table.setColumnWidth(col, 120)

    def updateWorkTable(self, start_time, end_time, work_type):
        current_date = self.parent.date_edit.date().toString('yyyy-MM-dd')
    
        for row in range(self.work_table.rowCount()):
            if self.work_table.item(row, 0).text() == current_date:
                # Clear previous entries for this row
                for col in range(2, self.work_table.columnCount()):
                    self.work_table.setItem(row, col, QTableWidgetItem(""))
            
                if work_type == "Sima munkanap":
                    self.work_table.setItem(row, 2, QTableWidgetItem(start_time))
                    self.work_table.setItem(row, 3, QTableWidgetItem(end_time))
                    # Calculate hours
                    start = QTime.fromString(start_time, "HH:mm")
                    end = QTime.fromString(end_time, "HH:mm")
                    hours = round(start.secsTo(end) / 3600.0, 2)
                    self.work_table.setItem(row, 4, QTableWidgetItem(str(hours)))
                
                elif work_type == "Műhely nap":
                    self.work_table.setItem(row, 5, QTableWidgetItem(start_time))
                    self.work_table.setItem(row, 6, QTableWidgetItem(end_time))
                    # Calculate hours
                    start = QTime.fromString(start_time, "HH:mm")
                    end = QTime.fromString(end_time, "HH:mm")
                    hours = round(start.secsTo(end) / 3600.0, 2)
                    self.work_table.setItem(row, 7, QTableWidgetItem(str(hours)))
                
                elif work_type == "Szabadság":
                    self.work_table.setItem(row, 8, QTableWidgetItem("1"))
                
                elif work_type == "Betegszabadság (TP)":
                    self.work_table.setItem(row, 9, QTableWidgetItem("1"))
            
                # Ensure all cells are center-aligned
                for col in range(self.work_table.columnCount()):
                    item = self.work_table.item(row, col)
                    if item:
                        item.setTextAlignment(Qt.AlignCenter)
                break

    def saveWorkHours(self):
        if not self.parent.driver_combo.currentText():
            QMessageBox.warning(self.parent, "Hiba", "Válasszon sofőrt!")
            return False

        try:
            current_driver = self.parent.driver_combo.currentText()
            month_dir = os.path.join('driver_records', current_driver, 
                                   f"{datetime.now().year}_{datetime.now().month:02d}")
            os.makedirs(month_dir, exist_ok=True)
    
            excel_path = os.path.join(month_dir, 'munkaora_nyilvantartas.xlsx')
    
            # Meglévő Excel betöltése vagy új létrehozása
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
            else:
                wb = Workbook()
            ws = wb.active
            ws.title = "Munkaórák"

            # Aktuális dátum és adatok mentése
            current_date = self.parent.date_edit.date().toString('yyyy-MM-dd')
            start_time = self.parent.start_time.time().toString('HH:mm')
            end_time = self.parent.end_time.time().toString('HH:mm')
            work_type = self.parent.type_combo.currentText()

            # Adatok frissítése a munkalapon
            self.updateWorkTable(start_time, end_time, work_type)

            # Fejléc cellák formázása
            headers = [
                "Dátum", "Nap", 
                "Sima Munkanap\nKezdés", "Sima Munkanap\nVégzés", "Ledolgozott\nÓrák",
                "Műhely\nKezdés", "Műhely\nVégzés", "Műhely\nÓrák",
                "Szabadság", "Betegszabadság\n(TP)"
            ]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border
                cell.alignment = Alignment(wrap_text=True)  # Szöveg tördelése

            # Óra számítás hozzáadása
            for row in range(2, ws.max_row + 1):
                start_time = ws.cell(row=row, column=3).value  # Kezdés oszlop
                end_time = ws.cell(row=row, column=4).value    # Végzés oszlop
                if start_time and end_time:
                    try:
                        start = datetime.strptime(start_time, '%H:%M')
                        end = datetime.strptime(end_time, '%H:%M')
                        hours = (end - start).seconds / 3600
                        ws.cell(row=row, column=5, value=round(hours, 2))  # Ledolgozott órák
                    except:
                        pass

            # Excel mentése
            for row in range(self.work_table.rowCount()):
                for col in range(self.work_table.columnCount()):
                    item = self.work_table.item(row, col)
                    if item:
                        ws.cell(row=row+2, column=col+1, value=item.text())

            # Oszlopok szélességének beállítása
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(excel_path)
            QMessageBox.information(self.parent, "Siker", "Munkaórák mentve!")
            return True

        except Exception as e:
            QMessageBox.critical(self.parent, "Hiba", f"Mentési hiba: {str(e)}")
            return False

    def loadWorkHours(self, current_driver):
        try:
            month_dir = os.path.join('driver_records', current_driver, 
                                   f"{datetime.now().year}_{datetime.now().month:02d}")
            excel_path = os.path.join(month_dir, 'munkaora_nyilvantartas.xlsx')
            
            if not os.path.exists(excel_path):
                return False

            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active

            # Clear existing data except dates and days
            for row in range(self.work_table.rowCount()):
                for col in range(2, self.work_table.columnCount()):
                    self.work_table.setItem(row, col, None)

            # Load data
            for excel_row in range(2, ws.max_row + 1):
                date = ws.cell(row=excel_row, column=1).value
                if not date:
                    continue
                
                for table_row in range(self.work_table.rowCount()):
                    table_date = self.work_table.item(table_row, 0)
                    if table_date and table_date.text() == str(date):
                        for col in range(2, self.work_table.columnCount()):
                            value = ws.cell(row=excel_row, column=col+1).value
                            if value is not None:
                                item = QTableWidgetItem(str(value))
                                item.setTextAlignment(Qt.AlignCenter)
                                self.work_table.setItem(table_row, col, item)
                        break

            return True

        except Exception as e:
            QMessageBox.warning(self.parent, "Hiba", f"Betöltési hiba: {str(e)}")
            return False

    def saveTransportData(self):
        if not self.parent.driver_combo.currentText():
            QMessageBox.warning(self.parent, "Hiba", "Válasszon sofőrt!")
            return False

        try:
            current_driver = self.parent.driver_combo.currentText()
            month_dir = os.path.join('driver_records', current_driver, 
                                   f"{datetime.now().year}_{datetime.now().month:02d}")
            os.makedirs(month_dir, exist_ok=True)
    
            excel_path = os.path.join(month_dir, 'fuvar_adatok.xlsx')
    
            # Meglévő Excel betöltése vagy új létrehozása
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
            else:
                wb = Workbook()
            ws = wb.active
            ws.title = "Fuvar Adatok"

            # Fejléc cellák formázása
            headers = [
                "Dátum", "Fuvar ID", "Indulás", "Érkezés", "Távolság", "Üzemanyag"
            ]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border
                cell.alignment = Alignment(wrap_text=True)  # Szöveg tördelése

            # Excel mentése
            for row in range(self.transport_table.rowCount()):
                for col in range(self.transport_table.columnCount()):
                    item = self.transport_table.item(row, col)
                    if item:
                        ws.cell(row=row+2, column=col+1, value=item.text())

            # Oszlopok szélességének beállítása
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            wb.save(excel_path)
            QMessageBox.information(self.parent, "Siker", "Fuvar adatok mentve!")
            return True

        except Exception as e:
            QMessageBox.critical(self.parent, "Hiba", f"Mentési hiba: {str(e)}")
            return False

    def loadTransportData(self, current_driver):
        try:
            month_dir = os.path.join('driver_records', current_driver, 
                                   f"{datetime.now().year}_{datetime.now().month:02d}")
            excel_path = os.path.join(month_dir, 'fuvar_adatok.xlsx')
            
            if not os.path.exists(excel_path):
                return False

            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active

            # Clear existing data
            for row in range(self.transport_table.rowCount()):
                for col in range(self.transport_table.columnCount()):
                    self.transport_table.setItem(row, col, None)

            # Load data
            for excel_row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    value = ws.cell(row=excel_row, column=col).value
                    if value is not None:
                        item = QTableWidgetItem(str(value))
                        item.setTextAlignment(Qt.AlignCenter)
                        self.transport_table.setItem(excel_row-2, col-1, item)

            return True

        except Exception as e:
            QMessageBox.warning(self.parent, "Hiba", f"Betöltési hiba: {str(e)}")
            return False
        self.setup_headers()

    def setup_transport_table(self, table):
        self.transport_table = table
        self.setup_transport_headers()

    def setup_headers(self):
        headers = [
            "Dátum", "Nap", 
            "Sima Munkanap\nKezdés", "Sima Munkanap\nVégzés", "Ledolgozott\nÓrák",
            "Műhely\nKezdés", "Műhely\nVégzés", "Műhely\nÓrák",
            "Szabadság", "Betegszabadság\n(TP)"
        ]
        self.work_table.setColumnCount(len(headers))
        self.work_table.setHorizontalHeaderLabels(headers)
    
        # Oszlopok szélességének beállítása
        for col in range(self.work_table.columnCount()):
            self.work_table.setColumnWidth(col, 120)

    def updateWorkTable(self, start_time, end_time, work_type):
        current_date = self.parent.date_edit.date().toString('yyyy-MM-dd')
    
        for row in range(self.work_table.rowCount()):
            if self.work_table.item(row, 0).text() == current_date:
                # Clear previous entries for this row
                for col in range(2, self.work_table.columnCount()):
                    self.work_table.setItem(row, col, QTableWidgetItem(""))
            
                if work_type == "Sima munkanap":
                    self.work_table.setItem(row, 2, QTableWidgetItem(start_time))
                    self.work_table.setItem(row, 3, QTableWidgetItem(end_time))
                    # Calculate hours
                    start = QTime.fromString(start_time, "HH:mm")
                    end = QTime.fromString(end_time, "HH:mm")
                    hours = round(start.secsTo(end) / 3600.0, 2)
                    self.work_table.setItem(row, 4, QTableWidgetItem(str(hours)))
                
                elif work_type == "Műhely nap":
                    self.work_table.setItem(row, 5, QTableWidgetItem(start_time))
                    self.work_table.setItem(row, 6, QTableWidgetItem(end_time))
                    # Calculate hours
                    start = QTime.fromString(start_time, "HH:mm")
                    end = QTime.fromString(end_time, "HH:mm")
                    hours = round(start.secsTo(end) / 3600.0, 2)
                    self.work_table.setItem(row, 7, QTableWidgetItem(str(hours)))
                
                elif work_type == "Szabadság":
                    self.work_table.setItem(row, 8, QTableWidgetItem("1"))
                
                elif work_type == "Betegszabadság (TP)":
                    self.work_table.setItem(row, 9, QTableWidgetItem("1"))
            
                # Ensure all cells are center-aligned
                for col in range(self.work_table.columnCount()):
                    item = self.work_table.item(row, col)
                    if item:
                        item.setTextAlignment(Qt.AlignCenter)
                break

    def saveWorkHours(self):
        if not self.parent.driver_combo.currentText():
            QMessageBox.warning(self.parent, "Hiba", "Válasszon sofőrt!")
            return False

        try:
            current_driver = self.parent.driver_combo.currentText()
            month_dir = os.path.join('driver_records', current_driver, 
                                   f"{datetime.now().year}_{datetime.now().month:02d}")
            os.makedirs(month_dir, exist_ok=True)
    
            excel_path = os.path.join(month_dir, 'munkaora_nyilvantartas.xlsx')
    
            # Meglévő Excel betöltése vagy új létrehozása
            if os.path.exists(excel_path):
                wb = load_workbook(excel_path)
            else:
                wb = Workbook()
            ws = wb.active
            ws.title = "Munkaórák"

            # Aktuális dátum és adatok mentése
            current_date = self.parent.date_edit.date().toString('yyyy-MM-dd')
            start_time = self.parent.start_time.time().toString('HH:mm')
            end_time = self.parent.end_time.time().toString('HH:mm')
            work_type = self.parent.type_combo.currentText()

            # Adatok frissítése a munkalapon
            self.updateWorkTable(start_time, end_time, work_type)

            # Fejléc cellák formázása
            headers = [
                "Dátum", "Nap", 
                "Sima Munkanap\nKezdés", "Sima Munkanap\nVégzés", "Ledolgozott\nÓrák",
                "Műhely\nKezdés", "Műhely\nVégzés", "Műhely\nÓrák",
                "Szabadság", "Betegszabadság\n(TP)"
            ]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            header_alignment = Alignment(horizontal="center", vertical="center")
            header_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = header_border

            # Óra számítás hozzáadása
            for row in range(2, ws.max_row + 1):
                start_time = ws.cell(row=row, column=3).value  # Kezdés oszlop
                end_time = ws.cell(row=row, column=4).value    # Végzés oszlop
                if start_time and end_time:
                    try:
                        start = datetime.strptime(start_time, '%H:%M')
                        end = datetime.strptime(end_time, '%H:%M')
                        hours = (end - start).seconds / 3600
                        ws.cell(row=row, column=5, value=round(hours, 2))  # Ledolgozott órák
                    except:
                        pass

            # Excel mentése
            for row in range(self.work_table.rowCount()):
                for col in range(self.work_table.columnCount()):
                    item = self.work_table.item(row, col)
                    if item:
                        ws.cell(row=row+2, column=col+1, value=item.text())

            wb.save(excel_path)
            QMessageBox.information(self.parent, "Siker", "Munkaórák mentve!")
            return True

        except Exception as e:
            QMessageBox.critical(self.parent, "Hiba", f"Mentési hiba: {str(e)}")
            return False

    def loadWorkHours(self, current_driver):
        try:
            month_dir = os.path.join('driver_records', current_driver, 
                                   f"{datetime.now().year}_{datetime.now().month:02d}")
            excel_path = os.path.join(month_dir, 'munkaora_nyilvantartas.xlsx')
            
            if not os.path.exists(excel_path):
                return False

            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active

            # Clear existing data except dates and days
            for row in range(self.work_table.rowCount()):
                for col in range(2, self.work_table.columnCount()):
                    self.work_table.setItem(row, col, None)

            # Load data
            for excel_row in range(2, ws.max_row + 1):
                date = ws.cell(row=excel_row, column=1).value
                if not date:
                    continue
                
                for table_row in range(self.work_table.rowCount()):
                    table_date = self.work_table.item(table_row, 0)
                    if table_date and table_date.text() == str(date):
                        for col in range(2, self.work_table.columnCount()):
                            value = ws.cell(row=excel_row, column=col+1).value
                            if value is not None:
                                item = QTableWidgetItem(str(value))
                                item.setTextAlignment(Qt.AlignCenter)
                                self.work_table.setItem(table_row, col, item)
                        break

            return True

        except Exception as e:
            QMessageBox.warning(self.parent, "Hiba", f"Betöltési hiba: {str(e)}")
            return False