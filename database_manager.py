# -*- coding: utf-8 -*-
from PySide6.QtCore import Qt, QDate
from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QWidget, QFormLayout,
    QLineEdit, QSpinBox, QPushButton, QHBoxLayout,
    QTableWidget, QTabWidget, QTableWidgetItem, QMenuBar,
    QMessageBox, QComboBox, QDateEdit, QCheckBox, QGridLayout,
    QFrame, QLabel, QFileDialog, QDoubleSpinBox
)
from PySide6.QtCore import Qt, QDate
from PySide6.QtWidgets import QMainWindow
from datetime import datetime
import os
from openpyxl import Workbook
from typing import Dict, Any
 
from address_manager import AddressManager
import sys
sys.path.append('.')
from database_handler import DatabaseHandler
from vacation_manager import VacationManager


class DatabaseManager(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumSize(1024, 768)
        self.db = DatabaseHandler()
        self.vacation_manager = VacationManager(self.db)
        self.setupDatabase()
        self.initUI()

    def showError(self, title, message):
        QMessageBox.critical(self, title, f"Hiba: {message}")

    def initUI(self):
        self.setWindowTitle("Törzsadat Kezelő")
        layout = QVBoxLayout()
        
        tabs = QTabWidget()
        tabs.addTab(self.createFactoriesTab(), "Gyárak")
        tabs.addTab(self.createDriversTab(), "Sofőrök")
        tabs.addTab(self.createVehiclesTab(), "Gépjárművek")
        tabs.addTab(self.createVacationTab(), "Szabadság")
        tabs.addTab(self.createFuelTab(), "Üzemanyag")
        
        layout.addWidget(tabs)
        self.setLayout(layout)
        self.setupConnections()

    def setupConnections(self):
        # Drivers tab
        self.driver_table.itemClicked.connect(self.onDriverSelected)
        self.add_driver_btn.clicked.connect(self.addDriver)
        self.save_driver_btn.clicked.connect(self.saveDriverChanges)
        self.delete_driver_btn.clicked.connect(self.deleteDriver)

        # Vehicles tab
        self.vehicles_table.itemClicked.connect(self.onVehicleSelected)
        self.add_vehicle_btn.clicked.connect(self.addVehicle)
        self.save_vehicle_btn.clicked.connect(self.saveVehicleChanges)
        self.delete_vehicle_btn.clicked.connect(self.deleteVehicle)

    def setupDatabase(self):
       try:
           # FK ellenőrzés kikapcsolása
           self.db.execute_query("PRAGMA foreign_keys=OFF")
       
           # Táblák létrehozása
           self.db.execute_query('''
               CREATE TABLE IF NOT EXISTS vehicles (
                   id INTEGER PRIMARY KEY,
                   plate_number TEXT NOT NULL,
                   type TEXT,
                   brand TEXT,
                   model TEXT,
                   year_of_manufacture INTEGER,
                   chassis_number TEXT,
                   engine_number TEXT, 
                   engine_type TEXT,
                   fuel_type TEXT,
                   max_weight INTEGER,
                   own_weight INTEGER,
                   payload_capacity INTEGER,
                   seats INTEGER,
                   technical_review_date TEXT,
                   tachograph_type TEXT,
                   tachograph_calibration_date TEXT,
                   fire_extinguisher_expiry TEXT
               )
           ''')

           self.db.execute_query('''
               CREATE TABLE IF NOT EXISTS factories (
                   id INTEGER PRIMARY KEY,
                   name TEXT NOT NULL,
                   created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
               )
           ''')

           self.db.execute_query('''
               CREATE TABLE IF NOT EXISTS factory_zone_prices (
                   id INTEGER PRIMARY KEY,
                   factory_id INTEGER,
                   zone_name TEXT, 
                   price INTEGER DEFAULT 0,
                   FOREIGN KEY (factory_id) REFERENCES factories(id)
               )
           ''')

           self.db.execute_query('''
               CREATE TABLE IF NOT EXISTS factory_waiting_fees (
                   id INTEGER PRIMARY KEY,
                   factory_id INTEGER NOT NULL,
                   price_per_15_min INTEGER DEFAULT 0,
                   FOREIGN KEY (factory_id) REFERENCES factories(id)
               )
           ''')

           self.db.execute_query('''
               CREATE TABLE IF NOT EXISTS drivers (
                   id INTEGER PRIMARY KEY,
                   name TEXT NOT NULL,
                   birth_date TEXT,
                   birth_place TEXT,
                   address TEXT,
                   mothers_name TEXT,
                   tax_number TEXT,
                   social_security_number TEXT,
                   bank_name TEXT,
                   bank_account TEXT,
                   drivers_license_number TEXT,
                   drivers_license_expiry TEXT,
                   vacation_days INTEGER DEFAULT 29,
                   used_vacation_days INTEGER DEFAULT 0
               )
           ''')

           self.db.execute_query('''
               CREATE TABLE IF NOT EXISTS vacation_allowance (
                   id INTEGER PRIMARY KEY,
                   year INTEGER NOT NULL,
                   total_days INTEGER DEFAULT 29,
                   used_days INTEGER DEFAULT 0,
                   UNIQUE(year)
               )
           ''')

           self.db.execute_query('''
               CREATE TABLE IF NOT EXISTS fuel_consumption (
                   id INTEGER PRIMARY KEY,
                   vehicle_id INTEGER,
                   date TEXT,
                   odometer_reading INTEGER,
                   fuel_amount REAL,
                   fuel_price REAL,
                   total_cost REAL,
                   location TEXT,
                   full_tank BOOLEAN,
                   avg_consumption REAL,
                   FOREIGN KEY (vehicle_id) REFERENCES vehicles(id)
               )
           ''')

           self.db.execute_query('''
               CREATE TABLE IF NOT EXISTS addresses (
                   id INTEGER PRIMARY KEY,
                   address TEXT NOT NULL UNIQUE,
                   price INTEGER DEFAULT 0
               )
           ''')

           self.db.execute_query('''
               CREATE TABLE IF NOT EXISTS deliveries (
                   id INTEGER PRIMARY KEY,
                   delivery_date TEXT NOT NULL,
                   driver_id INTEGER,
                   vehicle_id INTEGER,
                   factory_id INTEGER,
                   zone_id INTEGER,
                   address_id INTEGER,
                   delivery_number TEXT NOT NULL,
                   amount REAL,
                   status TEXT DEFAULT 'pending',
                   created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                   FOREIGN KEY (driver_id) REFERENCES drivers(id),
                   FOREIGN KEY (vehicle_id) REFERENCES vehicles(id),
                   FOREIGN KEY (factory_id) REFERENCES factories(id),
                   FOREIGN KEY (zone_id) REFERENCES factory_zone_prices(id),
                   FOREIGN KEY (address_id) REFERENCES addresses(id)
               )
           ''')

           # FK ellenőrzés visszakapcsolása
           self.db.execute_query("PRAGMA foreign_keys=ON")

       except Exception as e:
           print(f"Database initialization error: {str(e)}")
           raise

    def saveDriverChanges(self):
        selected = self.driver_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Figyelmeztetés", "Válasszon sofőrt!")
            return
        
        driver_id = int(self.driver_table.item(selected[0].row(), 0).text())
        driver_data = self._collectDriverData()
        driver_data['id'] = driver_id
    
        try:
            self.db.insert_record('drivers', driver_data)
            self.loadDrivers()
            QMessageBox.information(self, "Siker", "Változások mentve!")
        except Exception as e:
            self.showError("Mentési hiba", str(e))

    def loadDrivers(self):
        try:
            results = self.db.execute_query("""
                SELECT id, name, birth_date, birth_place, address, 
                       mothers_name, tax_number, social_security_number,
                       drivers_license_number, bank_name, bank_account
                FROM drivers 
                ORDER BY name""")
        
            self.driver_table.setRowCount(len(results))
            for row, driver in enumerate(results):
                fields = ['id', 'name', 'birth_date', 'birth_place', 'address',
                         'mothers_name', 'tax_number', 'social_security_number', 
                         'drivers_license_number', 'bank_name', 'bank_account']
            
                for col, field in enumerate(fields):
                    item = QTableWidgetItem(str(driver[field] if driver[field] else ""))
                    item.setTextAlignment(Qt.AlignCenter)
                    self.driver_table.setItem(row, col, item)
        except Exception as e:
            self.showError("Betöltési hiba", str(e))

    def deleteDriver(self):
        selected = self.driver_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Figyelmeztetés", "Válasszon sofőrt!")
            return
        
        driver_id = int(self.driver_table.item(selected[0].row(), 0).text())
    
        if QMessageBox.question(self, 'Megerősítés', 'Biztosan törli a sofőrt?',
                              QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            try:
                self.db.execute_query("DELETE FROM drivers WHERE id = ?", (driver_id,))
                self.loadDrivers()
                QMessageBox.information(self, "Siker", "Sofőr törölve!")
            except Exception as e:
                self.showError("Törlési hiba", str(e))

    def checkDataDirectories(self):
        base_dir = 'driver_records'
        if not os.path.exists(base_dir):
            os.makedirs(base_dir)
        
        current_driver = self.driver_combo.currentText()
        if current_driver:
            month_dir = os.path.join(base_dir, current_driver,
                                   f"{datetime.now().year}_{datetime.now().month:02d}")
            if not os.path.exists(month_dir):
                os.makedirs(month_dir)
            
            return month_dir
        return None

    def saveData(self):
        month_dir = self.checkDataDirectories()
        if not month_dir:
            return False
        
        try:
            # Save work hours
            work_path = os.path.join(month_dir, 'munkaora_nyilvantartas.xlsx')
            self.work_hours_manager.saveWorkHours(work_path)
        
            # Save delivery data
            delivery_path = os.path.join(month_dir, 'fuvar_nyilvantartas.xlsx')
            self.delivery_manager.saveDeliveryData(delivery_path)
        
            return True
        except Exception as e:
            QMessageBox.critical(self, "Hiba", f"Mentési hiba: {str(e)}")
            return False

    def createDriversTab(self):
            widget = QWidget()
            layout = QVBoxLayout()
        
            form_layout = QFormLayout()
        
            # Input mezők inicializálása
            self.driver_name = QLineEdit()
            self.birth_date = QDateEdit()
            self.birth_place = QLineEdit()
            self.address = QLineEdit()
            self.mothers_name = QLineEdit()
            self.tax_number = QLineEdit()
            self.social_security_number = QLineEdit()
            self.drivers_license_number = QLineEdit()
            self.drivers_license_expiry = QDateEdit()
            self.bank_name = QLineEdit()
            self.bank_account = QLineEdit()
        
            # Form feltöltése
            form_layout.addRow("Név:", self.driver_name)
            form_layout.addRow("Születési idő:", self.birth_date)
            form_layout.addRow("Születési hely:", self.birth_place)
            form_layout.addRow("Lakcím:", self.address)
            form_layout.addRow("Anyja neve:", self.mothers_name)
            form_layout.addRow("Adószám:", self.tax_number)
            form_layout.addRow("TAJ szám:", self.social_security_number)
            form_layout.addRow("Jogosítvány száma:", self.drivers_license_number)
            form_layout.addRow("Jogosítvány lejárata:", self.drivers_license_expiry)
            form_layout.addRow("Bank neve:", self.bank_name)
            form_layout.addRow("Bankszámlaszám:", self.bank_account)
        
            # Gombok
            btn_layout = QHBoxLayout()
            self.add_driver_btn = QPushButton("Hozzáadás")
            self.save_driver_btn = QPushButton("Mentés")
            self.delete_driver_btn = QPushButton("Törlés")
        
            btn_layout.addWidget(self.add_driver_btn)
            btn_layout.addWidget(self.save_driver_btn)
            btn_layout.addWidget(self.delete_driver_btn)
        
            # Táblázat
            self.driver_table = QTableWidget()
            self.driver_table.setColumnCount(11)
            self.driver_table.setHorizontalHeaderLabels([
                "ID", "Név", "Születési idő", "Születési hely", "Lakcím",
                "Anyja neve", "Adószám", "TAJ szám", "Jogosítvány száma",
                "Bank neve", "Bankszámlaszám"
            ])
        
            layout.addLayout(form_layout)
            layout.addLayout(btn_layout)
            layout.addWidget(self.driver_table)
        
            widget.setLayout(layout)
            return widget

    def onDriverSelected(self, item):
        try:
            row = item.row()
            self.driver_name.setText(self.driver_table.item(row, 1).text() if self.driver_table.item(row, 1) else "")
            self.birth_date.setDate(QDate.fromString(self.driver_table.item(row, 2).text(), 'yyyy-MM-dd') if self.driver_table.item(row, 2) else QDate.currentDate())
            self.birth_place.setText(self.driver_table.item(row, 3).text() if self.driver_table.item(row, 3) else "")
            self.address.setText(self.driver_table.item(row, 4).text() if self.driver_table.item(row, 4) else "")
            self.mothers_name.setText(self.driver_table.item(row, 5).text() if self.driver_table.item(row, 5) else "")
            self.tax_number.setText(self.driver_table.item(row, 6).text() if self.driver_table.item(row, 6) else "")
            self.social_security_number.setText(self.driver_table.item(row, 7).text() if self.driver_table.item(row, 7) else "")
            self.drivers_license_number.setText(self.driver_table.item(row, 8).text() if self.driver_table.item(row, 8) else "")
            self.drivers_license_expiry.setDate(QDate.fromString(self.driver_table.item(row, 9).text(), 'yyyy-MM-dd') if self.driver_table.item(row, 9) else QDate.currentDate())
            self.bank_name.setText(self.driver_table.item(row, 10).text() if self.driver_table.item(row, 10) else "")
            self.bank_account.setText(self.driver_table.item(row, 11).text() if self.driver_table.item(row, 11) else "")
        except Exception as e:
            QMessageBox.critical(self, "Hiba", f"Hiba történt az adatok betöltése során: {str(e)}")

    def createVehiclesTab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        
        form_layout = QFormLayout()
        
        # Alapadatok
        self.plate_number = QLineEdit()
        self.vehicle_type = QLineEdit()
        self.brand = QLineEdit()
        self.model = QLineEdit()
        self.year_of_manufacture = QSpinBox()
        self.year_of_manufacture.setRange(1900, QDate.currentDate().year())
        
        # Azonosítók
        self.chassis_number = QLineEdit()
        self.engine_number = QLineEdit()
        self.engine_type = QLineEdit()
        self.fuel_type = QComboBox()
        self.fuel_type.addItems(["Dízel", "Benzin", "Elektromos", "Hibrid"])
        
        # Műszaki adatok
        self.max_weight = QSpinBox()
        self.max_weight.setRange(0, 50000)
        self.max_weight.setSuffix(" kg")
        self.own_weight = QSpinBox()
        self.own_weight.setRange(0, 50000)
        self.own_weight.setSuffix(" kg")
        self.payload_capacity = QSpinBox()
        self.payload_capacity.setRange(0, 50000)
        self.payload_capacity.setSuffix(" kg")
        self.seats = QSpinBox()
        self.seats.setRange(1, 50)
        
        # Dátumok és speciális adatok
        self.technical_review_date = QDateEdit()
        self.technical_review_date.setCalendarPopup(True)
        self.tachograph_type = QLineEdit()
        self.tachograph_calibration_date = QDateEdit()
        self.tachograph_calibration_date.setCalendarPopup(True)
        self.fire_extinguisher_expiry = QDateEdit()
        self.fire_extinguisher_expiry.setCalendarPopup(True)
        
        # Mezők hozzáadása az űrlaphoz
        form_layout.addRow("Rendszám:", self.plate_number)
        form_layout.addRow("Típus:", self.vehicle_type)
        form_layout.addRow("Márka:", self.brand)
        form_layout.addRow("Model:", self.model)
        form_layout.addRow("Gyártási év:", self.year_of_manufacture)
        form_layout.addRow("Alvázszám:", self.chassis_number)
        form_layout.addRow("Motorszám:", self.engine_number)
        form_layout.addRow("Motor típus:", self.engine_type)
        form_layout.addRow("Üzemanyag:", self.fuel_type)
        form_layout.addRow("Össztömeg:", self.max_weight)
        form_layout.addRow("Saját tömeg:", self.own_weight)
        form_layout.addRow("Hasznos terhelés:", self.payload_capacity)
        form_layout.addRow("Ülések száma:", self.seats)
        form_layout.addRow("Műszaki vizsga:", self.technical_review_date)
        form_layout.addRow("Tachográf típus:", self.tachograph_type)
        form_layout.addRow("Tachográf hitelesítés:", self.tachograph_calibration_date)
        form_layout.addRow("Tűzoltó készülék lejárat:", self.fire_extinguisher_expiry)
        
        # createVehiclesTab-ben:
        btn_layout = QHBoxLayout()
        self.add_vehicle_btn = QPushButton("Hozzáadás")
        self.save_vehicle_btn = QPushButton("Mentés")
        self.delete_vehicle_btn = QPushButton("Törlés")

        self.add_vehicle_btn.clicked.connect(self.addVehicle)
        self.save_vehicle_btn.clicked.connect(self.saveVehicleChanges)
        self.delete_vehicle_btn.clicked.connect(self.deleteVehicle)

        btn_layout.addWidget(self.add_vehicle_btn)
        btn_layout.addWidget(self.save_vehicle_btn)
        btn_layout.addWidget(self.delete_vehicle_btn)
        
        # Táblázat
        self.vehicles_table = QTableWidget()
        self.vehicles_table.setColumnCount(17)
        self.vehicles_table.setHorizontalHeaderLabels([
            "ID", "Rendszám", "Típus", "Márka", "Model", "Gyártási év",
            "Alvázszám", "Motorszám", "Motor típus", "Üzemanyag",
            "Össztömeg", "Saját tömeg", "Hasznos terhelés", "Ülések száma",
            "Műszaki vizsga", "Tachográf típus", "Tachográf hitelesítés"
        ])
        self.vehicles_table.itemClicked.connect(self.onVehicleSelected)
        
        layout.addLayout(form_layout)
        layout.addLayout(btn_layout)
        layout.addWidget(self.vehicles_table)
        
        widget.setLayout(layout)
        self.loadVehicles()
        return widget

    def saveVehicleChanges(self):
        """Jármű módosításainak mentése"""
        selected = self.vehicles_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Figyelmeztetés", "Válasszon járművet!")
            return
            
        vehicle_id = int(self.vehicles_table.item(selected[0].row(), 0).text())
        
        try:
            data = {
                'plate_number': self.plate_number.text().strip(),
                'type': self.vehicle_type.text().strip(),
                'brand': self.brand.text().strip(),
                'model': self.model.text().strip(),
                'year_of_manufacture': self.year_of_manufacture.value(),
                'chassis_number': self.chassis_number.text().strip(),
                'engine_number': self.engine_number.text().strip(),
                'engine_type': self.engine_type.text().strip(),
                'fuel_type': self.fuel_type.currentText(),
                'max_weight': self.max_weight.value(),
                'own_weight': self.own_weight.value(),
                'payload_capacity': self.payload_capacity.value(),
                'seats': self.seats.value(),
                'technical_review_date': self.technical_review_date.date().toString('yyyy-MM-dd'),
                'tachograph': self.tachograph_type.text().strip(),
                'tachograph_calibration_date': self.tachograph_calibration_date.date().toString('yyyy-MM-dd'),
                'fire_extinguisher_expiry': self.fire_extinguisher_expiry.date().toString('yyyy-MM-dd')
            }
            
            # SQL update lekérdezés készítése
            columns = [f"{key} = ?" for key in data.keys()]
            query = f"UPDATE vehicles SET {', '.join(columns)} WHERE id = ?"
            params = list(data.values()) + [vehicle_id]
            
            self.db.execute_query(query, tuple(params))
            self.loadVehicles()
            QMessageBox.information(self, "Siker", "Jármű adatai sikeresen frissítve!")
            
        except Exception as e:
            QMessageBox.critical(self, "Hiba", f"Mentési hiba: {str(e)}")

    def loadVehicles(self):
       try:
           # Debug információ
           results = self.db.execute_query("""
               SELECT id, plate_number, type, brand, model, 
                      year_of_manufacture, chassis_number, engine_number,
                      engine_type, fuel_type, max_weight, own_weight,
                      payload_capacity, seats, technical_review_date,
                      tachograph_type, tachograph_calibration_date, 
                      fire_extinguisher_expiry
               FROM vehicles
               ORDER BY plate_number
           """)

           self.vehicles_table.setRowCount(len(results))
       
           # Minden járműhöz végigmegyünk az oszlopokon
           for row, vehicle in enumerate(results):
               # ID
               self.vehicles_table.setItem(row, 0, QTableWidgetItem(str(vehicle['id'])))
           
               # Rendszám
               self.vehicles_table.setItem(row, 1, QTableWidgetItem(vehicle['plate_number']))
           
               # Típus
               self.vehicles_table.setItem(row, 2, QTableWidgetItem(vehicle['type']))
           
               # Márka
               self.vehicles_table.setItem(row, 3, QTableWidgetItem(vehicle['brand']))
           
               # Modell
               self.vehicles_table.setItem(row, 4, QTableWidgetItem(vehicle['model']))
           
               # Gyártási év 
               self.vehicles_table.setItem(row, 5, QTableWidgetItem(str(vehicle['year_of_manufacture'])))
           
               # Alvázszám
               self.vehicles_table.setItem(row, 6, QTableWidgetItem(vehicle['chassis_number']))
           
               # Motorszám
               self.vehicles_table.setItem(row, 7, QTableWidgetItem(vehicle['engine_number']))
           
               # Motor típus
               self.vehicles_table.setItem(row, 8, QTableWidgetItem(vehicle['engine_type']))
           
               # Üzemanyag
               self.vehicles_table.setItem(row, 9, QTableWidgetItem(vehicle['fuel_type']))
           
               # Össztömeg
               self.vehicles_table.setItem(row, 10, QTableWidgetItem(f"{vehicle['max_weight']} kg"))
           
               # Saját tömeg
               self.vehicles_table.setItem(row, 11, QTableWidgetItem(f"{vehicle['own_weight']} kg"))
           
               # Hasznos terhelés
               self.vehicles_table.setItem(row, 12, QTableWidgetItem(f"{vehicle['payload_capacity']} kg"))
           
               # Ülések száma
               self.vehicles_table.setItem(row, 13, QTableWidgetItem(str(vehicle['seats'])))
           
               # Műszaki vizsga
               self.vehicles_table.setItem(row, 14, QTableWidgetItem(vehicle['technical_review_date']))
           
               # Tachográf típus
               self.vehicles_table.setItem(row, 15, QTableWidgetItem(vehicle['tachograph_type']))
           
               # Tachográf hitelesítés
               self.vehicles_table.setItem(row, 16, QTableWidgetItem(vehicle['tachograph_calibration_date']))
           
               # Minden cella középre igazítása
               for col in range(self.vehicles_table.columnCount()):
                   item = self.vehicles_table.item(row, col)
                   if item:
                       item.setTextAlignment(Qt.AlignCenter)
       
           # Oszlopok méretezése
           self.vehicles_table.resizeColumnsToContents()
       
           for col in range(self.vehicles_table.columnCount()):
               width = self.vehicles_table.columnWidth(col)
               if width < 100:  # Minimum szélesség
                   self.vehicles_table.setColumnWidth(col, 100)
               elif width > 200:  # Maximum szélesség
                   self.vehicles_table.setColumnWidth(col, 200)

       except Exception as e:
           print(f"Betöltési hiba: {str(e)}")  # Debug információ
           QMessageBox.critical(self, "Hiba", f"Járművek betöltési hiba: {str(e)}")

    def deleteVehicle(self):
        """Jármű törlése"""
        selected = self.vehicles_table.selectedItems()
        if not selected:
            QMessageBox.warning(self, "Figyelmeztetés", "Válasszon járművet!")
            return
        
        vehicle_id = int(self.vehicles_table.item(selected[0].row(), 0).text())
        plate = self.vehicles_table.item(selected[0].row(), 1).text()
        
        reply = QMessageBox.question(self, 'Megerősítés', 
                                   f'Biztosan törli a {plate} rendszámú járművet?',
                                   QMessageBox.Yes | QMessageBox.No)
                                   
        if reply == QMessageBox.Yes:
            try:
                self.db.execute_query("DELETE FROM vehicles WHERE id = ?", (vehicle_id,))
                self.loadVehicles()
                self.clearVehicleFields()
                QMessageBox.information(self, "Siker", "Jármű sikeresen törölve!")
            except Exception as e:
                QMessageBox.critical(self, "Hiba", f"Törlési hiba: {str(e)}")

    def clearVehicleFields(self):
        self.plate_number.clear()
        self.vehicle_type.clear()
        self.brand.clear()
        self.model.clear()
        self.year_of_manufacture.setValue(QDate.currentDate().year())
        self.chassis_number.clear()
        self.engine_number.clear()
        self.engine_type.clear()
        self.fuel_type.setCurrentIndex(0)
        self.max_weight.setValue(0)
        self.own_weight.setValue(0)
        self.payload_capacity.setValue(0)
        self.seats.setValue(2)
        self.technical_review_date.setDate(QDate.currentDate())
        self.tachograph_type.clear()
        self.tachograph_calibration_date.setDate(QDate.currentDate())
        self.fire_extinguisher_expiry.setDate(QDate.currentDate())

    def addDriver(self):
        if not self._validateDriverInput():
            return
        
        try:
            driver_data = self._collectDriverData()
            self.db.execute_query("""
                INSERT INTO drivers (
                    name, birth_date, birth_place, address, 
                    mothers_name, tax_number, social_security_number,
                    drivers_license_number, drivers_license_expiry,
                    bank_name, bank_account
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                driver_data['name'], driver_data['birth_date'],
                driver_data['birth_place'], driver_data['address'],
                driver_data['mothers_name'], driver_data['tax_number'],
                driver_data['social_security_number'],
                driver_data['drivers_license_number'],
                driver_data['drivers_license_expiry'],
                driver_data['bank_name'], driver_data['bank_account']
            ))
            
            self.loadDrivers()
            self.clearDriverFields()
            QMessageBox.information(self, "Siker", "Sofőr sikeresen hozzáadva!")
            
        except Exception as e:
            QMessageBox.critical(self, "Hiba", f"Mentési hiba: {str(e)}")

    def loadFactories(self):
        try:
            factories = self.db.execute_query("""
                SELECT f.id, f.name, fw.price_per_15_min 
                FROM factories f
                LEFT JOIN factory_waiting_fees fw ON f.id = fw.factory_id
                ORDER BY f.name
            """)
            self.factory_table.setRowCount(len(factories))
            for row, factory in enumerate(factories):
                self.factory_table.setItem(row, 0, QTableWidgetItem(str(factory['id'])))
                self.factory_table.setItem(row, 1, QTableWidgetItem(factory['name']))
                self.factory_table.setItem(row, 2, QTableWidgetItem(f"{factory['price_per_15_min']} Ft"))
        except Exception as e:
            self.showError("Betöltési hiba", str(e))

    def deleteFactory(self):
        try:
            selected = self.factory_table.selectedItems()
            if not selected:
                QMessageBox.warning(self, "Figyelmeztetés", "Válasszon gyárat!")
                return
            
            factory_id = int(self.factory_table.item(selected[0].row(), 0).text())
            factory_name = self.factory_table.item(selected[0].row(), 1).text()
        
            reply = QMessageBox.question(self, 'Megerősítés', 
                                       f'Biztosan törli a {factory_name} gyárat?',
                                       QMessageBox.Yes | QMessageBox.No)
                                   
            if reply == QMessageBox.Yes:
                self.db.execute_query("DELETE FROM factory_waiting_fees WHERE factory_id = ?", (factory_id,))
                self.db.execute_query("DELETE FROM factory_zone_prices WHERE factory_id = ?", (factory_id,))
                self.db.execute_query("DELETE FROM factories WHERE id = ?", (factory_id,))
            
                self.loadFactories()
                self.factory_name.clear()
                self.waiting_fee.setValue(0)
                QMessageBox.information(self, "Siker", "Gyár sikeresen törölve!")
            
        except Exception as e:
            QMessageBox.critical(self, "Hiba", f"Törlési hiba: {str(e)}")

    def _validateDriverInput(self) -> bool:
        """Sofőr adatok validálása"""
        if not self.driver_name.text().strip():
            QMessageBox.warning(self, "Figyelmeztetés", "A név megadása kötelező!")
            return False
        return True

    def _collectDriverData(self) -> Dict[str, Any]:
        """Sofőr adatok összegyűjtése"""
        return {
            'name': self.driver_name.text().strip(),
            'birth_date': self.birth_date.date().toString('yyyy-MM-dd'),
            'birth_place': self.birth_place.text().strip(),
            'address': self.address.text().strip(),
            'mothers_name': self.mothers_name.text().strip(),
            'tax_number': self.tax_number.text().strip(),
            'social_security_number': self.social_security_number.text().strip(),
            'drivers_license_number': self.drivers_license_number.text().strip(),
            'drivers_license_expiry': self.drivers_license_expiry.date().toString('yyyy-MM-dd'),
            'bank_name': self.bank_name.text().strip(),
            'bank_account': self.bank_account.text().strip(),
            
        }

    def clearDriverFields(self):
        self.driver_name.clear()
        self.birth_place.clear()
        self.address.clear()
        self.mothers_name.clear()
        self.tax_number.clear()
        self.social_security_number.clear()
        self.drivers_license_number.clear()
        self.bank_name.clear()
        self.bank_account.clear()
        self.birth_date.setDate(QDate.currentDate())
        self.drivers_license_expiry.setDate(QDate.currentDate())

    def onVehicleSelected(self, item):
       try:
           row = item.row()
           # Minden mező óvatos beállítása
           self.plate_number.setText(self.vehicles_table.item(row, 1).text() if self.vehicles_table.item(row, 1) else "")
           self.vehicle_type.setText(self.vehicles_table.item(row, 2).text() if self.vehicles_table.item(row, 2) else "")
           self.brand.setText(self.vehicles_table.item(row, 3).text() if self.vehicles_table.item(row, 3) else "")
           self.model.setText(self.vehicles_table.item(row, 4).text() if self.vehicles_table.item(row, 4) else "")
           self.year_of_manufacture.setValue(int(self.vehicles_table.item(row, 5).text()) if self.vehicles_table.item(row, 5) and self.vehicles_table.item(row, 5).text().isdigit() else 1900)
           self.chassis_number.setText(self.vehicles_table.item(row, 6).text() if self.vehicles_table.item(row, 6) else "")
           self.engine_number.setText(self.vehicles_table.item(row, 7).text() if self.vehicles_table.item(row, 7) else "")
           self.engine_type.setText(self.vehicles_table.item(row, 8).text() if self.vehicles_table.item(row, 8) else "")
           self.fuel_type.setCurrentText(self.vehicles_table.item(row, 9).text() if self.vehicles_table.item(row, 9) else "Dízel")
           self.max_weight.setValue(int(self.vehicles_table.item(row, 10).text()) if self.vehicles_table.item(row, 10) and self.vehicles_table.item(row, 10).text().isdigit() else 0)
           self.own_weight.setValue(int(self.vehicles_table.item(row, 11).text()) if self.vehicles_table.item(row, 11) and self.vehicles_table.item(row, 11).text().isdigit() else 0)
           self.payload_capacity.setValue(int(self.vehicles_table.item(row, 12).text()) if self.vehicles_table.item(row, 12) and self.vehicles_table.item(row, 12).text().isdigit() else 0)
           self.seats.setValue(int(self.vehicles_table.item(row, 13).text()) if self.vehicles_table.item(row, 13) and self.vehicles_table.item(row, 13).text().isdigit() else 2)
       
           # Dátumok beállítása
           date_fields = [
               (self.technical_review_date, 14),
               (self.tachograph_calibration_date, 16),
               (self.fire_extinguisher_expiry, 17)
           ]
           for date_field, col in date_fields:
               if self.vehicles_table.item(row, col) and self.vehicles_table.item(row, col).text():
                   date_field.setDate(QDate.fromString(self.vehicles_table.item(row, col).text(), 'yyyy-MM-dd'))
               else:
                   date_field.setDate(QDate.currentDate())
       except Exception as e:
           QMessageBox.critical(self, "Hiba", f"Hiba történt az adatok betöltése során: {str(e)}")

    def addVehicle(self) -> None:
        if not self._validateVehicleInput():  # Első ellenőrzés
            return

        try:
            vehicle_data = self._collectVehicleData()
            # Még egy ellenőrzés az adatbázis művelet előtt
            if not vehicle_data['plate_number']:
                QMessageBox.warning(self, "Figyelmeztetés", "A rendszám megadása kötelező!")
                return
            
            self.db.insert_record('vehicles', vehicle_data)
            self.loadVehicles()  # Frissítjük a táblázatot
            QMessageBox.information(self, "Siker", "Gépjármű sikeresen hozzáadva!")
            self.clearVehicleFields()
        
        except Exception as e:
            QMessageBox.critical(self, "Hiba", f"Mentési hiba: {str(e)}")

    def _validateVehicleInput(self) -> bool:
        """Jármű adatok validálása"""
        if not self.plate_number.text().strip():
            QMessageBox.warning(self, "Figyelmeztetés", "A rendszám megadása kötelező!")
            return False
        return True

    def _collectVehicleData(self) -> Dict[str, Any]:
        """Jármű adatok összegyűjtése"""
        return {
            'plate_number': self.plate_number.text().strip(),
            'type': self.vehicle_type.text().strip(),
            'brand': self.brand.text().strip(),
            'model': self.model.text().strip(),
            'year_of_manufacture': self.year_of_manufacture.value(),
            'chassis_number': self.chassis_number.text().strip(),
            'engine_number': self.engine_number.text().strip(),
            'engine_type': self.engine_type.text().strip(),
            'fuel_type': self.fuel_type.currentText(),
            'max_weight': self.max_weight.value(),
            'own_weight': self.own_weight.value(),
            'payload_capacity': self.payload_capacity.value(),
            'seats': self.seats.value(),
            'technical_review_date': self.technical_review_date.date().toString('yyyy-MM-dd'),
            'tachograph_type': self.tachograph_type.text().strip(),
            'tachograph_calibration_date': self.tachograph_calibration_date.date().toString('yyyy-MM-dd'),
            'fire_extinguisher_expiry': self.fire_extinguisher_expiry.date().toString('yyyy-MM-dd')
        }

    def createVacationTab(self):
        widget = QWidget()
        layout = QVBoxLayout()
    
        form_layout = QFormLayout()
        self.year_combo = QComboBox()
        current_year = QDate.currentDate().year()
        self.year_combo.addItems([str(year) for year in range(current_year-1, current_year+2)])
        self.year_combo.setCurrentText(str(current_year))
    
        self.vacation_days = QSpinBox()
        self.vacation_days.setRange(0, 100)
        self.vacation_days.setSingleStep(1)
    
        form_layout.addRow("Év:", self.year_combo)
        form_layout.addRow("Szabadság napok:", self.vacation_days)
    
        btn_layout = QHBoxLayout()
        save_btn = QPushButton("Mentés")
        delete_btn = QPushButton("Törlés")
        save_btn.clicked.connect(self.saveVacationDays)
        delete_btn.clicked.connect(self.deleteVacationDays)
        btn_layout.addWidget(save_btn)
        btn_layout.addWidget(delete_btn)
    
        self.vacation_table = QTableWidget()
        self.vacation_table.setColumnCount(3)
        self.vacation_table.setHorizontalHeaderLabels(["Év", "Összes nap", "Felhasznált napok"])
    
        layout.addLayout(form_layout)
        layout.addLayout(btn_layout)
        layout.addWidget(self.vacation_table)
        widget.setLayout(layout)
    
        self.loadVacationData()
        return widget

    def createFactoriesTab(self):
        widget = QWidget()
        layout = QVBoxLayout()
    
        # Gyár alapadatok bevitele
        form_layout = QFormLayout()
        self.factory_name = QLineEdit()
        self.waiting_fee = QSpinBox()
        self.waiting_fee.setRange(0, 100000)
        self.waiting_fee.setSingleStep(500)
        form_layout.addRow("Gyár neve:", self.factory_name)
        form_layout.addRow("Állásidő díj (Ft/15 perc):", self.waiting_fee)
    
        # Övezeti díj hozzáadása szekció
        zone_layout = QHBoxLayout()
        self.zone_combo = QComboBox()
        self.zone_combo.addItems([f"Övezet {i}-{i+5}" for i in range(0, 50, 5)])
        self.zone_price = QSpinBox()
        self.zone_price.setRange(0, 100000)
        self.zone_price.setSingleStep(500)
    
        zone_layout.addWidget(self.zone_combo)
        zone_layout.addWidget(self.zone_price)
        form_layout.addRow("Övezeti díj:", zone_layout)
    
        # Gombok
        btn_layout = QHBoxLayout()
        add_factory_btn = QPushButton("Gyár hozzáadása")
        add_factory_btn.clicked.connect(self.addFactory)
        add_zone_btn = QPushButton("Övezeti díj hozzáadása")
        add_zone_btn.clicked.connect(self.addZonePrice)
        save_changes_btn = QPushButton("Változások mentése")
        save_changes_btn.clicked.connect(self.saveFactoryChanges)
        delete_factory_btn = QPushButton("Gyár törlése")
        delete_factory_btn.clicked.connect(self.deleteFactory)
    
        btn_layout.addWidget(add_factory_btn)
        btn_layout.addWidget(add_zone_btn)
        btn_layout.addWidget(save_changes_btn)
        btn_layout.addWidget(delete_factory_btn)
    
        # Gyárak táblázat
        self.factory_table = QTableWidget()
        self.factory_table.setColumnCount(3)
        self.factory_table.setHorizontalHeaderLabels(["ID", "Név", "Állásidő díj"])
        self.factory_table.itemClicked.connect(self.onFactorySelected)
    
        # Övezeti díjak táblázat
        self.zone_prices_table = QTableWidget()
        self.zone_prices_table.setColumnCount(3)
        self.zone_prices_table.setHorizontalHeaderLabels(["Övezet", "Díj (Ft)", ""])
    
        layout.addLayout(form_layout)
        layout.addLayout(btn_layout)
        layout.addWidget(self.factory_table)
        layout.addWidget(self.zone_prices_table)
    
        widget.setLayout(layout)
        self.loadFactories()
        return widget

    def onFactorySelected(self, item):
        try:
            row = item.row()
            factory_id = int(self.factory_table.item(row, 0).text())
        
            # Gyár nevének és állásidő díjának beállítása
            self.factory_name.setText(self.factory_table.item(row, 1).text())
            self.waiting_fee.setValue(int(self.factory_table.item(row, 2).text().replace(" Ft", "")))
        
            # Övezeti árak betöltése
            self.loadZonePrices(factory_id)
        
        except Exception as e:
            QMessageBox.critical(self, "Hiba", f"Hiba az adatok betöltésekor: {str(e)}")

    def addZonePrice(self):
        selected_items = self.factory_table.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "Hiba", "Először válasszon ki egy gyárat!")
            return
        
        factory_id = int(self.factory_table.item(selected_items[0].row(), 0).text())
        zone = self.zone_combo.currentText()
        price = self.zone_price.value()
    
        try:
            self.db.execute_query(
                "UPDATE factory_zone_prices SET price = ? WHERE factory_id = ? AND zone_name = ?",
                (price, factory_id, zone)
            )
        
            self.loadZonePrices(factory_id)
            self.zone_price.setValue(0)
        
            QMessageBox.information(self, "Siker", "Övezeti díj sikeresen módosítva!")
        
        except Exception as e:
            QMessageBox.critical(self, "Hiba", f"Hiba történt a mentés során: {str(e)}")

    def openAddressManager(self):
        address_manager = AddressManager(self.parent())
        address_manager.exec()
        # Címkezelő bezárása után frissítsük a főablak címlistáját
        self.parent().loadAddresses()

    def addFactory(self):
        try:
            name = self.factory_name.text()
            if not name:
                QMessageBox.warning(self, "Figyelmeztetés", "A gyár nevét kötelező megadni!")
                return

            # Gyár beszúrása
            factory_id = self.db.execute_query(
                "INSERT INTO factories (name) VALUES (?)",
                (name,)
            )
        
            # Állásidő díj beszúrása
            self.db.execute_query(
                "INSERT INTO factory_waiting_fees (factory_id, price_per_15_min) VALUES (?, ?)",
                (factory_id, self.waiting_fee.value())
            )
        
            # Övezeti díjak inicializálása
            for i in range(0, 50, 5):
                zone_name = f"Övezet {i}-{i+5}"
                self.db.execute_query(
                    "INSERT INTO factory_zone_prices (factory_id, zone_name, price) VALUES (?, ?, 0)",
                    (factory_id, zone_name)
                )

            self.loadFactories()
            self.factory_name.clear()
            self.waiting_fee.setValue(0)
            QMessageBox.information(self, "Siker", "Gyár sikeresen hozzáadva!")

        except Exception as e:
            QMessageBox.critical(self, "Hiba", f"Hiba történt a mentés során: {str(e)}")

    def deleteZonePrice(self, factory_id, zone):
        try:
            self.db.execute_query(
                "UPDATE factory_zone_prices SET price = 0 WHERE factory_id = ? AND zone_name = ?",
                (factory_id, zone)
            )
            self.loadZonePrices(factory_id)
        except Exception as e:
            self.showError("Törlési hiba", str(e))

    def saveFactoryChanges(self):
       selected_items = self.factory_table.selectedItems()
       if not selected_items:
           QMessageBox.warning(self, "Hiba", "Először válasszon ki egy gyárat!")
           return
       
       try:
           row = selected_items[0].row()
           factory_id = int(self.factory_table.item(row, 0).text())
           waiting_fee = int(self.waiting_fee.value())
       
           self.db.execute_query(
               "UPDATE factory_waiting_fees SET price_per_15_min = ? WHERE factory_id = ?",
               (waiting_fee, factory_id)
           )
       
           QMessageBox.information(self, "Siker", "Változások mentve!")
       
       except Exception as e:
           self.showError("Mentési hiba", str(e))

    def deleteAddress(self):
        try:
            selected_items = self.address_table.selectedItems()
            if not selected_items:
                QMessageBox.warning(self, "Figyelmeztetés", "Kérem válasszon ki egy címet!")
                return
        
            address_id = int(self.address_table.item(selected_items[0].row(), 0).text())
            address = self.address_table.item(selected_items[0].row(), 1).text()
    
            reply = QMessageBox.question(self, 'Megerősítés', 
                                       f'Biztosan törölni szeretné a következő címet: {address}?',
                                       QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
    
            if reply == QMessageBox.Yes:
                self.db.execute_query("DELETE FROM addresses WHERE id = ?", (address_id,))
                self.loadAddresses()
                QMessageBox.information(self, "Siker", "Cím sikeresen törölve!")
        
        except Exception as e:
            self.showError("Törlési hiba", str(e))

    def loadAddresses(self):
        try:
            addresses = self.db.execute_query("SELECT id, address, price FROM addresses ORDER BY address")
            self.address_table.setRowCount(len(addresses))
            for row, addr in enumerate(addresses):
                self.address_table.setItem(row, 0, QTableWidgetItem(str(addr['id'])))
                self.address_table.setItem(row, 1, QTableWidgetItem(addr['address']))
                self.address_table.setItem(row, 2, QTableWidgetItem(f"{addr['price']:,} Ft"))
        except Exception as e:
            self.showError("Betöltési hiba", str(e))

    def onWorkTypeChanged(self, work_type):
        if work_type == "Szabadság":
            self.vacation_frame.show()
            self.updateVacationDisplay()
        else:
            self.vacation_frame.hide()

    def saveVacationDays(self):
        try:
            year = int(self.year_combo.currentText())
            total_days = self.vacation_days.value()
    
            self.db.execute_query("""
                INSERT OR REPLACE INTO vacation_allowance (year, total_days, used_days)
                VALUES (?, ?, COALESCE((SELECT used_days FROM vacation_allowance WHERE year = ?), 0))
            """, (year, total_days, year))
    
            self.loadVacationData()
            QMessageBox.information(self, "Siker", "Szabadság keret sikeresen mentve!")
    
            if isinstance(self.parent(), QMainWindow):
                self.parent().updateVacationDisplay()
    
        except Exception as e:
            QMessageBox.critical(self, "Hiba", f"Hiba történt a mentés során: {str(e)}")

    def deleteVacationDays(self):
        try:
            if not self.vacation_table.selectedItems():
                QMessageBox.warning(self, "Figyelmeztetés", "Kérem válasszon ki egy sort!")
                return
            
            year = int(self.vacation_table.item(self.vacation_table.selectedItems()[0].row(), 0).text())
            if QMessageBox.question(self, 'Megerősítés', f'Biztosan törli a {year} évi szabadság keretet?',
                                  QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
                self.db.execute_query("DELETE FROM vacation_allowance WHERE year = ?", (year,))
                self.loadVacationData()
            
                if isinstance(self.parent(), QMainWindow):
                    self.parent().updateVacationDisplay()
                
        except Exception as e:
            self.showError("Törlési hiba", str(e))

    def loadVehiclesForFuel(self): 
        try:
            vehicles = self.db.execute_query("SELECT id, plate_number FROM vehicles ORDER BY plate_number")
            self.fuel_vehicle_combo.clear()
            for vehicle in vehicles:
                self.fuel_vehicle_combo.addItem(vehicle['plate_number'], vehicle['id'])
        except Exception as e:
            self.showError("Betöltési hiba", str(e))

    def loadZonePrices(self, factory_id):
        try:
            prices = self.db.execute_query("""
                SELECT zone_name, price 
                FROM factory_zone_prices 
                WHERE factory_id = ?
                ORDER BY zone_name""", 
                (factory_id,))
            
            self.zone_prices_table.setRowCount(len(prices))
            for row, price in enumerate(prices):
                self.zone_prices_table.setItem(row, 0, QTableWidgetItem(price['zone_name']))
                self.zone_prices_table.setItem(row, 1, QTableWidgetItem(f"{price['price']:,} Ft"))
            
                delete_btn = QPushButton("Törlés")
                delete_btn.clicked.connect(lambda c, z=price['zone_name']: 
                                         self.deleteZonePrice(factory_id, z))
                self.zone_prices_table.setCellWidget(row, 2, delete_btn)
            
        except Exception as e:
            self.showError("Betöltési hiba", str(e))

    def loadVacationData(self):
        try:
            results = self.db.execute_query(
                "SELECT year, total_days, used_days FROM vacation_allowance ORDER BY year DESC"
            )
        
            self.vacation_table.setRowCount(len(results))
            for row, record in enumerate(results):
                self.vacation_table.setItem(row, 0, QTableWidgetItem(str(record['year'])))
                self.vacation_table.setItem(row, 1, QTableWidgetItem(str(record['total_days'])))
                self.vacation_table.setItem(row, 2, QTableWidgetItem(str(record['used_days'])))
            
        except Exception as e:
            self.showError("Betöltési hiba", str(e))

    def setupFuelDatabase(self):
        self.db.execute_query('''
            CREATE TABLE IF NOT EXISTS fuel_consumption (
                id INTEGER PRIMARY KEY,
                vehicle_id INTEGER,
                date TEXT,
                odometer_reading INTEGER,
                fuel_amount REAL,
                fuel_price REAL,
                total_cost REAL,
                location TEXT,
                full_tank BOOLEAN,
                avg_consumption REAL,
                FOREIGN KEY (vehicle_id) REFERENCES vehicles(id)
            )
        ''')

    def createFuelTab(self):
        widget = QWidget()
        layout = QVBoxLayout()
    
        # Adatbeviteli űrlap
        form_layout = QFormLayout()
    
        self.fuel_vehicle_combo = QComboBox()
        self.loadVehiclesForFuel()
    
        self.fuel_date = QDateEdit()
        self.fuel_date.setCalendarPopup(True)
        self.fuel_date.setDate(QDate.currentDate())
    
        # Itt hozzuk létre az odometer QSpinBox-ot
        self.odometer = QSpinBox()
        self.odometer.setRange(0, 9999999)  # Megfelelő tartomány beállítása
        self.odometer.setSuffix(" km")      # Mértékegység hozzáadása
    
        self.fuel_amount = QDoubleSpinBox()
        self.fuel_amount.setRange(0, 1000)
        self.fuel_amount.setDecimals(2)
    
        self.fuel_price = QDoubleSpinBox()
        self.fuel_price.setRange(0, 10000)
        self.fuel_price.setDecimals(2)
    
        self.total_cost = QLineEdit()
        self.total_cost.setReadOnly(True)
    
        self.location = QLineEdit()
    
        self.full_tank = QCheckBox("Tele tank")
    
        # Mezők hozzáadása
        form_layout.addRow("Jármű:", self.fuel_vehicle_combo)
        form_layout.addRow("Dátum:", self.fuel_date)
        form_layout.addRow("Kilométeróra állás:", self.odometer)
        form_layout.addRow("Üzemanyag mennyiség (L):", self.fuel_amount)
        form_layout.addRow("Üzemanyag ár (Ft/L):", self.fuel_price)
        form_layout.addRow("Összköltség (Ft):", self.total_cost)
        form_layout.addRow("Helyszín:", self.location)
        form_layout.addRow("", self.full_tank)
    
        # Események
        self.fuel_amount.valueChanged.connect(self.calculateTotalCost)
        self.fuel_price.valueChanged.connect(self.calculateTotalCost)
    
        # Gombok
        btn_layout = QHBoxLayout()
        add_btn = QPushButton("Hozzáadás")
        add_btn.clicked.connect(self.addFuelRecord)
        delete_btn = QPushButton("Törlés")
        delete_btn.clicked.connect(self.deleteFuelRecord)
    
        btn_layout.addWidget(add_btn)
        btn_layout.addWidget(delete_btn)
    
        # Táblázat
        self.fuel_table = QTableWidget()
        self.fuel_table.setColumnCount(10)
        self.fuel_table.setHorizontalHeaderLabels([
            "ID", "Jármű", "Dátum", "Kilométeróra", "Mennyiség (L)",
            "Ár (Ft/L)", "Összköltség (Ft)", "Helyszín", "Tele tank",
            "Átlagfogyasztás (L/100km)"
        ])
    
        layout.addLayout(form_layout)
        layout.addLayout(btn_layout)
        layout.addWidget(self.fuel_table)
    
        widget.setLayout(layout)
        self.loadFuelRecords()
        return widget

    def calculateFuelConsumption(self, previous_record, current_record):
        try:
            distance = current_record[3] - previous_record[3]  # kilométeróra különbség
            fuel_amount = current_record[4]  # üzemanyag mennyiség
        
            if distance > 0 and fuel_amount > 0:
                consumption = (fuel_amount / distance) * 100
                return round(consumption, 2)
            return 0
        except:
            return 0

    def addFuelRecord(self):
        try:
            vehicle_id = self.fuel_vehicle_combo.currentData()
            new_odometer = int(self.odometer.value())
            new_fuel_amount = float(self.fuel_amount.value())
            is_full_tank = self.full_tank.isChecked()

            # Előző tankolás lekérdezése
            last_record = self.db.execute_query('''
                SELECT odometer_reading, fuel_amount 
                FROM fuel_consumption 
                WHERE vehicle_id = ? AND full_tank = TRUE
                ORDER BY date DESC, odometer_reading DESC 
                LIMIT 1
            ''', (vehicle_id,))

            # Átlagfogyasztás számítása
            avg_consumption = None
            if last_record and is_full_tank:
                prev_odometer = last_record[0]['odometer_reading']
                distance = new_odometer - prev_odometer
                if distance > 0:
                    avg_consumption = (new_fuel_amount / distance) * 100

            # Új rekord beszúrása
            self.db.execute_query('''
                INSERT INTO fuel_consumption (
                    vehicle_id, date, odometer_reading, fuel_amount,
                    fuel_price, total_cost, location, full_tank,
                    avg_consumption
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                vehicle_id,
                self.fuel_date.date().toString('yyyy-MM-dd'),
                new_odometer,
                new_fuel_amount,
                self.fuel_price.value(),
                float(self.total_cost.text().replace(' ', '')) if self.total_cost.text() else 0,
                self.location.text(),
                is_full_tank,
                avg_consumption
            ))

            self.loadFuelRecords()

            if avg_consumption is not None:
                QMessageBox.information(self, "Információ", 
                    f"Átlagfogyasztás: {avg_consumption:.2f} L/100km")
        
        except Exception as e:
            self.showError("Mentési hiba", str(e))

    def onVehicleChanged(self, index):
        try:
            vehicle_id = self.fuel_vehicle_combo.currentData()
            if vehicle_id is None:
                return
            
            previous_record = self.db.execute_query('''
                SELECT date
                FROM fuel_consumption 
                WHERE vehicle_id = ? 
                ORDER BY date DESC
                LIMIT 1
            ''', (vehicle_id,))
        
            if previous_record and previous_record[0]['date']:
                try:
                    last_date = QDate.fromString(previous_record[0]['date'], 'yyyy-MM-dd')
                    self.fuel_date.setDate(last_date)
                except:
                    self.fuel_date.setDate(QDate.currentDate())
        except Exception as e:
            QMessageBox.critical(self, "Hiba", f"Hiba történt: {str(e)}")

    def calculateTotalCost(self):
        amount = self.fuel_amount.value()
        price = self.fuel_price.value()
        total = amount * price
        self.total_cost.setText(f"{total:.2f}")

    def loadFuelRecords(self):
        try:
            results = self.db.execute_query('''
                SELECT 
                    f.id,
                    v.plate_number,
                    f.date,
                    f.odometer_reading,
                    f.fuel_amount,
                    f.fuel_price,
                    f.total_cost,
                    f.location,
                    f.full_tank,
                    f.avg_consumption
                FROM fuel_consumption f
                JOIN vehicles v ON f.vehicle_id = v.id
                ORDER BY f.date DESC, f.odometer_reading DESC
            ''')
        
            self.fuel_table.setRowCount(len(results))
            for row, record in enumerate(results):
                for col, value in enumerate([
                    str(record['id']),
                    record['plate_number'],
                    record['date'],
                    str(record['odometer_reading']),
                    f"{record['fuel_amount']:.1f}",
                    f"{record['fuel_price']:.2f}",
                    f"{record['total_cost']:.2f}",
                    record['location'],
                    "Igen" if record['full_tank'] else "Nem",
                    f"{record['avg_consumption']:.2f}" if record['avg_consumption'] else ""
                ]):
                    item = QTableWidgetItem(value)
                    item.setTextAlignment(Qt.AlignCenter)
                    self.fuel_table.setItem(row, col, item)
                
            self.fuel_table.resizeColumnsToContents()
        
        except Exception as e:
            self.showError("Betöltési hiba", str(e))

    def deleteFuelRecord(self):
        try:
            selected_items = self.fuel_table.selectedItems()
            if not selected_items:
                QMessageBox.warning(self, "Figyelmeztetés", "Kérem válasszon ki egy rekordot!")
                return
        
            record_id = int(self.fuel_table.item(selected_items[0].row(), 0).text())
        
            reply = QMessageBox.question(self, 'Megerősítés', 
                                       'Biztosan törli a kiválasztott rekordot?',
                                       QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                               
            if reply == QMessageBox.Yes:
                self.db.execute_query("DELETE FROM fuel_consumption WHERE id = ?", (record_id,))
                self.loadFuelRecords()
        
        except Exception as e:
            QMessageBox.critical(self, "Hiba", f"Hiba történt a törlés során: {str(e)}")

    def generateFuelReport(self, vehicle_id=None, start_date=None, end_date=None):
        try:
            query = '''
                SELECT 
                    v.plate_number,
                    COUNT(*) as tankolas_szam,
                    SUM(f.fuel_amount) as ossz_mennyiseg,
                    SUM(f.total_cost) as ossz_koltseg,
                    MAX(f.odometer_reading) - MIN(f.odometer_reading) as megtett_km,
                    CASE 
                        WHEN MAX(f.odometer_reading) - MIN(f.odometer_reading) > 0 
                        THEN (SUM(f.fuel_amount) / (MAX(f.odometer_reading) - MIN(f.odometer_reading))) * 100
                        ELSE 0 
                    END as atlag_fogyasztas
                FROM fuel_consumption f
                JOIN vehicles v ON f.vehicle_id = v.id
                WHERE 1=1
            '''
    
            params = []
            if vehicle_id:
                query += " AND v.id = ?"
                params.append(vehicle_id)
            if start_date:
                query += " AND f.date >= ?"
                params.append(start_date)
            if end_date:
                query += " AND f.date <= ?"
                params.append(end_date)
        
            query += " GROUP BY v.id, v.plate_number"
    
            return self.db.execute_query(query, params)
    
        except Exception as e:
            QMessageBox.critical(self, "Hiba", f"Hiba történt a riport generálása során: {str(e)}")
            return []

    def createBillingTab(self):
        widget = QWidget()
        layout = QVBoxLayout()

        # Filter frame with white background and visible text
        filter_frame = QFrame()
        filter_frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border: 1px solid #cccccc;
                border-radius: 5px;
                padding: 10px;
            }
            QLabel {
                color: black;
                font-weight: bold;
            }
            QComboBox, QDateEdit {
                color: black;
                background-color: white;
                border: 1px solid #cccccc;
                padding: 5px;
                min-width: 150px;
            }
        """)
    
        filter_layout = QGridLayout()

        # Date range
        self.date_from = QDateEdit()
        self.date_to = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_to.setCalendarPopup(True)
    
        filter_layout.addWidget(QLabel("Időszak:"), 0, 0)
        filter_layout.addWidget(self.date_from, 0, 1)
        filter_layout.addWidget(QLabel("-"), 0, 2)
        filter_layout.addWidget(self.date_to, 0, 3)

        # Combos
        self.billing_driver_combo = QComboBox()
        self.billing_factory_combo = QComboBox()
        self.billing_zone_combo = QComboBox()
    
        filter_layout.addWidget(QLabel("Sofőr:"), 1, 0)
        filter_layout.addWidget(self.billing_driver_combo, 1, 1)
        filter_layout.addWidget(QLabel("Gyár:"), 1, 2)
        filter_layout.addWidget(self.billing_factory_combo, 1, 3)
        filter_layout.addWidget(QLabel("Övezet:"), 1, 4)
        filter_layout.addWidget(self.billing_zone_combo, 1, 5)

        filter_frame.setLayout(filter_layout)
        layout.addWidget(filter_frame)

        # Table
        self.billing_table = QTableWidget()
        self.billing_table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                color: black;
                gridline-color: #cccccc;
            }
            QHeaderView::section {
                background-color: #f0f0f0;
                color: black;
                padding: 5px;
                border: 1px solid #cccccc;
            }
        """)
    
        layout.addWidget(self.billing_table)

        # Bottom buttons
        btn_layout = QHBoxLayout()
        for btn_text in ["Frissítés", "Excel export", "Számlázottnak jelöl"]:
            btn = QPushButton(btn_text)
            btn.setStyleSheet("""
                QPushButton {
                    background-color: #4a90e2;
                    color: white;
                    padding: 8px 16px;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #357abd;
                }
            """)
            btn_layout.addWidget(btn)
    
        layout.addLayout(btn_layout)
        widget.setLayout(layout)
    
        self.loadBillingData()
        return widget

    def loadBillingData(self):
        try:
            conditions = ["1=1"]
            params = []

            if self.date_from.date():
                conditions.append("d.delivery_date >= ?")
                params.append(self.date_from.date().toString('yyyy-MM-dd'))

            if self.date_to.date():
                conditions.append("d.delivery_date <= ?")
                params.append(self.date_to.date().toString('yyyy-MM-dd'))

            if self.billing_driver_combo.currentData() != -1:
                conditions.append("d.driver_id = ?")
                params.append(self.billing_driver_combo.currentData())

            if self.billing_factory_combo.currentData() != -1:
                conditions.append("d.factory_id = ?")
                params.append(self.billing_factory_combo.currentData())

            query = f"""
                SELECT 
                    d.delivery_date,
                    dr.name as driver_name,
                    f.name as factory_name,
                    z.zone_name,
                    d.delivery_number,
                    d.amount,
                    z.price as unit_price,
                    a.address,
                    d.status
                FROM deliveries d
                LEFT JOIN drivers dr ON d.driver_id = dr.id
                LEFT JOIN factories f ON d.factory_id = f.id
                LEFT JOIN zone_prices z ON d.zone_id = z.id
                LEFT JOIN addresses a ON d.address_id = a.id
                WHERE {' AND '.join(conditions)}
            """

            deliveries = self.db.execute_query(query, params)

            self.billing_table.setRowCount(len(deliveries))
            for row, delivery in enumerate(deliveries):
                for col, value in enumerate(delivery):
                    if col == 6:  # Összeg számítás
                        amount = delivery[5]
                        price = value
                        total = amount * price if amount and price else 0
                        item = QTableWidgetItem(f"{total:,.0f} Ft")
                    else:
                        item = QTableWidgetItem(str(value if value else ""))
                    item.setTextAlignment(Qt.AlignCenter)
                    self.billing_table.setItem(row, col, item)

        except Exception as e:
               self.showError("Adatok betöltése sikertelen", str(e))

    def refreshBillingItems(self):
       try:
           conditions = ["1=1"]
           params = []
       
           if self.date_from.date():
               conditions.append("d.delivery_date >= ?")
               params.append(self.date_from.date().toString('yyyy-MM-dd'))
       
           if self.date_to.date():
               conditions.append("d.delivery_date <= ?")
               params.append(self.date_to.date().toString('yyyy-MM-dd'))
       
           if self.billing_driver_combo.currentData() != -1:
               conditions.append("d.driver_id = ?")
               params.append(self.billing_driver_combo.currentData())
       
           if self.billing_factory_combo.currentData() != -1:
               conditions.append("d.factory_id = ?")
               params.append(self.billing_factory_combo.currentData())
       
           query = """
               SELECT 
                   d.delivery_date,
                   dr.name as driver_name,
                   f.name as factory_name,
                   z.zone_name,
                   COUNT(*) as delivery_count,
                   SUM(d.amount) as total_amount,
                   z.price as zone_price,
                   SUM(d.amount * z.price) as total_price,
                   d.status
               FROM deliveries d
               JOIN drivers dr ON d.driver_id = dr.id
               JOIN factories f ON d.factory_id = f.id
               JOIN factory_zone_prices z ON d.zone_id = z.id
               WHERE {}
               GROUP BY d.delivery_date, d.driver_id, d.factory_id, d.zone_id
               ORDER BY d.delivery_date DESC
           """.format(" AND ".join(conditions))

           rows = self.db.execute_query(query, params)
       
           self.billing_table.setRowCount(len(rows))
           for i, row in enumerate(rows):
               for j, value in enumerate(row):
                   item = QTableWidgetItem(str(value))
                   if isinstance(value, (int, float)) and j not in (0, 1, 2, 3, 8):
                       item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                       if j == 6:  # Egységár
                           item.setText(f"{value:,.0f} Ft")
                       elif j == 7:  # Összeg
                           item.setText(f"{value:,.0f} Ft")
                       else:  # Mennyiségek
                           item.setText(f"{value:,.1f}")
                   else:
                       item.setTextAlignment(Qt.AlignCenter)
                   self.billing_table.setItem(i, j, item)

           self.billing_table.resizeColumnsToContents()

       except Exception as e:
           self.showError("Adatok betöltése sikertelen", str(e))

    def exportBillingItems(self):
        try:
            filename, _ = QFileDialog.getSaveFileName(
                self, "Excel mentése", "", "Excel fájlok (*.xlsx)")
        
            if filename:
                wb = Workbook()
                ws = wb.active
            
                # Fejlécek
                headers = []
                for col in range(self.billing_table.columnCount()):
                    headers.append(self.billing_table.horizontalHeaderItem(col).text())
            
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
            
                # Adatok
                for row in range(self.billing_table.rowCount()):
                    for col in range(self.billing_table.columnCount()):
                        value = self.billing_table.item(row, col).text()
                        ws.cell(row=row+2, column=col+1, value=value)
            
                wb.save(filename)
                QMessageBox.information(self, "Siker", "Excel exportálás sikeres!")
            
        except Exception as e:
            QMessageBox.critical(self, "Hiba", f"Excel exportálási hiba: {str(e)}")

    def markItemsAsBilled(self):
       try:
           selected_rows = set(item.row() for item in self.billing_table.selectedItems())
           if not selected_rows:
               QMessageBox.warning(self, "Figyelmeztetés", "Válasszon ki tételeket!")
               return
       
           reply = QMessageBox.question(
               self, 'Megerősítés', 
               'Biztosan számlázottnak jelöli a kiválasztott tételeket?',
               QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
       
           if reply == QMessageBox.Yes:
               for row in selected_rows:
                   date = self.billing_table.item(row, 0).text()
                   driver_name = self.billing_table.item(row, 1).text()
                   factory_name = self.billing_table.item(row, 2).text()
                   zone_name = self.billing_table.item(row, 3).text()
               
                   self.db.execute_query("""
                       UPDATE deliveries 
                       SET status = 'billed'
                       WHERE delivery_date = ?
                       AND driver_id = (SELECT id FROM drivers WHERE name = ?)
                       AND factory_id = (SELECT id FROM factories WHERE name = ?)
                       AND zone_id = (SELECT id FROM zone_prices WHERE zone_name = ?)
                   """, (date, driver_name, factory_name, zone_name))
           
               self.refreshBillingItems()
               QMessageBox.information(self, "Siker", "Tételek sikeresen megjelölve!")
           
       except Exception as e:
           self.showError("Státusz módosítása sikertelen", str(e))