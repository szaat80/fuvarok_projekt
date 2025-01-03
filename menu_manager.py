# -*- coding: utf-8 -*-
from PySide6.QtWidgets import (QMenuBar, QMenu, QMessageBox, QDialog,
                              QInputDialog, QFileDialog, QTableWidgetItem)
from PySide6.QtPrintSupport import QPrintDialog, QPrinter
from PySide6.QtGui import QTextDocument, QTextCursor, QTextTableFormat
from PySide6.QtCore import Qt, QTimer
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
from database_manager import DatabaseManager
from address_manager import AddressManager

class MenuBar(QMenuBar):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.createFileMenu()
        self.createDatabaseMenu()
        self.createSettingsMenu()  # Új menüpont

    def createSettingsMenu(self):
        settingsMenu = self.addMenu("Beállítások")
        settings_action = settingsMenu.addAction("Beállítások")
        settings_action.triggered.connect(self.parent().showSettingsDialog)

    def saveWorkHours(self):
        if hasattr(self.parent(), 'work_hours_manager'):
            self.parent().work_hours_manager.saveWorkHours()

    def saveDelivery(self):
        if hasattr(self.parent(), 'delivery_manager'):
            self.parent().delivery_manager.saveDeliveryData()

    def openExcelFile(self):
        try:
            filename, _ = QFileDialog.getOpenFileName(
                self.parent(),
                "Excel fájl megnyitása",
                "",
                "Excel fájlok (*.xlsx *.xls)"
            )
            if filename:
                import os
                os.startfile(filename)
        except Exception as e:
            QMessageBox.warning(self.parent(), "Hiba", f"Excel megnyitási hiba: {str(e)}")

    def createFileMenu(self):
        fileMenu = self.addMenu("Fájl")
    
        excel_action = fileMenu.addAction("Excel megnyitása")
        excel_action.triggered.connect(self.openExcelFile)
    
        fileMenu.addSeparator()
        
        # Címkezelő hozzáadása
        address_action = fileMenu.addAction("Címek kezelése")
        address_action.triggered.connect(self.openAddressManager)
        
        fileMenu.addAction("Munkaórák mentése").triggered.connect(self.saveWorkHours)
        fileMenu.addAction("Fuvar mentése").triggered.connect(self.saveDelivery)
        fileMenu.addSeparator()
        fileMenu.addAction("Kilépés").triggered.connect(self.parent().close)

    def createDatabaseMenu(self):
        dbMenu = self.addMenu("Adatbázis")
        dbMenu.addAction("Törzsadatok kezelése").triggered.connect(self.openDatabaseManager)

    def openExcel(self, filename):
        try:
            import os
            if os.path.exists(filename):
                os.startfile(filename)
            else:
                QMessageBox.warning(self.parent(), "Figyelmeztetés", 
                                  f"{filename} nem található!")
        except Exception as e:
            QMessageBox.warning(self.parent(), "Hiba", 
                              f"Excel megnyitási hiba: {str(e)}")

    def loadWorkHoursData(self, driver):
        if driver:
            self.parent().work_hours_manager.loadWorkHours(driver)

    def loadDeliveryData(self, driver):
        if driver:
            self.parent().delivery_manager.loadDeliveryData(driver)

    def loadFromExcel(self, excel_path='munkaora_nyilvantartas.xlsx'):
        try:
            if not os.path.exists(excel_path):
                return
            
            wb = load_workbook(excel_path)
            ws = wb.active
            
            work_table = self.parent().work_table
            work_table.clearContents()
            
            for row in range(2, ws.max_row + 1):
                date = ws.cell(row=row, column=1).value
                if not date:
                    continue
                    
                for table_row in range(work_table.rowCount()):
                    table_date = work_table.item(table_row, 0)
                    if table_date and table_date.text() == str(date):
                        for col in range(work_table.columnCount()):
                            value = ws.cell(row=row, column=col+1).value
                            if value is not None:
                                item = QTableWidgetItem(str(value))
                                item.setTextAlignment(Qt.AlignCenter)
                                work_table.setItem(table_row, col, item)
                        break
            
            QMessageBox.information(self.parent(), "Siker", "Excel adatok betöltve!")
            
        except Exception as e:
            QMessageBox.warning(self.parent(), "Hiba", f"Excel betöltési hiba: {str(e)}")

    def openAddressManager(self):
        try:
            address_manager = AddressManager(self.parent())
            address_manager.exec()
            # Címkezelő bezárása után frissítsük a főablak címlistáját
            if hasattr(self.parent(), 'loadAddresses'):
                self.parent().loadAddresses()
        except Exception as e:
            QMessageBox.critical(self.parent(), "Hiba", f"Hiba a címkezelő megnyitásakor: {str(e)}")

    def openDatabaseManager(self):
        try:
            db_manager = DatabaseManager(self.parent())
            db_manager.exec()
            # DatabaseManager bezárása után frissítsük a főablak gyárlistáját
            if hasattr(self.parent(), 'loadFactories'):
                self.parent().loadFactories()
        except Exception as e:
            QMessageBox.critical(self.parent(), "Hiba", f"Hiba a törzsadat kezelő megnyitásakor: {str(e)}")

    def printData(self):
        try:
            printer = QPrinter()
            dialog = QPrintDialog(printer, self.parent())
            
            if dialog.exec() == QDialog.Accepted:
                document = QTextDocument()
                cursor = QTextCursor(document)
                
                # Table format
                table_format = QTextTableFormat()
                table_format.setBorder(1)
                table_format.setCellPadding(5)
                table_format.setAlignment(Qt.AlignCenter)
                
                # Work hours table
                work_table = self.parent().work_table
                cursor.insertText("Munkaóra nyilvántartás\n\n")
                table = cursor.insertTable(
                    work_table.rowCount() + 1,
                    work_table.columnCount(),
                    table_format
                )
                
                # Headers
                for col in range(work_table.columnCount()):
                    header = work_table.horizontalHeaderItem(col).text()
                    cell = table.cellAt(0, col)
                    cell_cursor = cell.firstCursorPosition()
                    cell_cursor.insertText(header)
                
                # Data
                for row in range(work_table.rowCount()):
                    for col in range(work_table.columnCount()):
                        item = work_table.item(row, col)
                        cell = table.cellAt(row + 1, col)
                        cell_cursor = cell.firstCursorPosition()
                        cell_cursor.insertText(item.text() if item else "")
                
                cursor.movePosition(QTextCursor.End)
                cursor.insertBlock()
                cursor.insertBlock()
                
                # Delivery table
                delivery_table = self.parent().delivery_table
                cursor.insertText("Fuvar nyilvántartás\n\n")
                table = cursor.insertTable(
                    delivery_table.rowCount() + 1,
                    delivery_table.columnCount(),
                    table_format
                )
                
                # Headers
                for col in range(delivery_table.columnCount()):
                    header = delivery_table.horizontalHeaderItem(col).text()
                    cell = table.cellAt(0, col)
                    cell_cursor = cell.firstCursorPosition()
                    cell_cursor.insertText(header)
                
                # Data
                for row in range(delivery_table.rowCount()):
                    for col in range(delivery_table.columnCount()):
                        item = delivery_table.item(row, col)
                        cell = table.cellAt(row + 1, col)
                        cell_cursor = cell.firstCursorPosition()
                        cell_cursor.insertText(item.text() if item else "")
                
                document.print_(printer)
                QMessageBox.information(self.parent(), "Siker", "Nyomtatás sikeres!")
                
        except Exception as e:
            QMessageBox.critical(self.parent(), "Hiba", f"Nyomtatási hiba: {str(e)}")