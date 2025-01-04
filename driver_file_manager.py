# -*- coding: utf-8 -*-
import os
import json
import shutil
from datetime import datetime
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

class DriverFileManager:
    def __init__(self, base_dir="sofor_nyilvantartas"):
        self.base_dir = base_dir
        self.conn = sqlite3.connect('fuvarok.db')
        self.setup_directory_structure()

    def setup_directory_structure(self):
        """Create base directory if it doesn't exist"""
        if not os.path.exists(self.base_dir):
            os.makedirs(self.base_dir)

    def get_drivers(self):
        """Fetch all drivers from the database"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT name FROM drivers")
        return [driver[0] for driver in cursor.fetchall()]

    def create_driver_folders(self):
        """Create folders for each driver with monthly subfolders"""
        drivers = self.get_drivers()
        current_year = datetime.now().year
        
        for driver in drivers:
            driver_dir = os.path.join(self.base_dir, driver)
            if not os.path.exists(driver_dir):
                os.makedirs(driver_dir)
            
            # Create monthly folders for current year
            for month in range(1, 13):
                month_dir = os.path.join(driver_dir, f"{current_year}_{month:02d}")
                if not os.path.exists(month_dir):
                    os.makedirs(month_dir)

    def organize_work_hours(self, current_driver=None, source_file='work_hours.json'):
        if current_driver is None:
            return
        month_dir = os.path.join(self.base_dir, current_driver, 
                                f"{datetime.now().year}_{datetime.now().month:02d}")
        os.makedirs(month_dir, exist_ok=True)
        excel_path = os.path.join(month_dir, 'work_hours.xlsx')
        
        with open(source_file, 'r', encoding='utf-8') as f:
            for line in f:
                try:
                    data = json.loads(line)
                    date = datetime.strptime(data['date'], '%Y-%m-%d')
                    driver = data.get('driver', 'Unknown')
                    
                    # Create monthly Excel file
                    month_dir = os.path.join(self.base_dir, driver, f"{date.year}_{date.month:02d}")
                    excel_path = os.path.join(month_dir, 'work_hours.xlsx')
                    
                    self.update_work_hours_excel(excel_path, data)
                except json.JSONDecodeError:
                    continue

    def organize_delivery_data(self, current_driver=None):
        if current_driver is None:
            return
        
        try:
            month_dir = os.path.join(self.base_dir, current_driver, 
                                    f"{datetime.now().year}_{datetime.now().month:02d}")
            os.makedirs(month_dir, exist_ok=True)
            excel_path = os.path.join(month_dir, 'fuvar_nyilvantartas.xlsx')
        
            wb = Workbook()
            ws = wb.active
            ws.title = "Fuvar adatok"
        
            # Add headers
            headers = ["Datum"]
            headers.extend([f"Övezet {i}-{i+5}" for i in range(0, 45, 5)])
        
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Save directly without using JSON
            wb.save(excel_path)
            return True
        
        except Exception as e:
            print(f"Error organizing delivery data: {str(e)}")
            return False

    def update_work_hours_excel(self, excel_path, data):
        """Update or create work hours Excel file"""
        wb = load_workbook(excel_path) if os.path.exists(excel_path) else Workbook()
        ws = wb.active if wb.active else wb.create_sheet()
        ws.title = "Work Hours"
        
        # Add headers if new file
        if ws.max_row == 1:
            headers = ["Date", "Start Time", "End Time", "Type", "Hours"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
        
        # Add new data
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1, value=data['date'])
        ws.cell(row=new_row, column=2, value=data.get('start_time', ''))
        ws.cell(row=new_row, column=3, value=data.get('end_time', ''))
        ws.cell(row=new_row, column=4, value=data.get('type', ''))
        ws.cell(row=new_row, column=5, value=data.get('hours', 0))
        
        wb.save(excel_path)

    def update_delivery_excel_fixed(excel_path, data):
        """Update or create delivery Excel file with duplicate prevention."""
        wb = load_workbook(excel_path) if os.path.exists(excel_path) else Workbook()
        ws = wb.active if wb.active else wb.create_sheet()
        ws.title = "Deliveries"

        # Add headers if new file
        headers = ["Date", "Factory", "Zone", "Address", "Delivery Number", "M3 Values"]
        if ws.max_row == 1:
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

        # Check for duplicate entries
        is_duplicate = False
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].value == data['date'] and row[4].value == data['delivery_number']:
                is_duplicate = True
                break

        if is_duplicate:
            print(f"Duplicate entry found for {data['delivery_number']} on {data['date']}. Skipping.")
            return

        # Add new data
        new_row = ws.max_row + 1
        ws.cell(row=new_row, column=1, value=data['date'])
        ws.cell(row=new_row, column=2, value=data.get('factory', '-'))
        ws.cell(row=new_row, column=3, value=data.get('km_range', '-'))
        ws.cell(row=new_row, column=4, value=data.get('address', '-'))
        ws.cell(row=new_row, column=5, value=data.get('delivery_number', '-'))

        # Handle different M3 value formats
        m3_values = data.get('m3_values', [])
        m3_str = ', '.join(map(str, m3_values)) if isinstance(m3_values, list) else str(m3_values)
        ws.cell(row=new_row, column=6, value=m3_str)

        wb.save(excel_path)


def main():
    manager = DriverFileManager()
    manager.create_driver_folders()
    manager.organize_work_hours()
    manager.organize_delivery_data()

if __name__ == "__main__":
    main()